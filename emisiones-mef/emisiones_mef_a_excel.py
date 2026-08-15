#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
emisiones_mef_a_excel.py
========================
Extrae los cuadros del PDF "Emisiones internas de bonos soberanos" (MEF,
2001-2026) y genera un Excel con formato uniforme.

Estrategia: tres motores de extracción complementarios (celdas de tabla,
alineación por coordenada vertical, y líneas de texto). Para cada año se usa
el motor cuya suma de Monto Colocado cuadra EXACTA contra la fila "Total"
impresa en el propio PDF. Los años sin cuadre exacto se marcan en la hoja
"Validación" con su diferencia.

Uso:
    pip install pdfplumber pandas openpyxl
    python emisiones_mef_a_excel.py <archivo.pdf> [salida.xlsx]
"""
import re
import sys

import pdfplumber
import pandas as pd

# ----------------------------------------------------------------------------
# Totales de referencia (fila "Total" de cada año, transcrita del PDF)
# ----------------------------------------------------------------------------
REF = {2001: 1_200_000_000, 2002: 733_200_000, 2003: 1_711_750_000,
       2004: 2_494_576_000, 2005: 6_643_629_000, 2006: 4_886_889_000,
       2007: 7_247_182_000, 2008: 1_227_000_000, 2009: 2_756_882_000,
       2010: 8_832_737_000, 2011: 1_220_660_000, 2012: 2_939_517_000,
       2013: 4_118_598_739, 2014: 11_435_817_466, 2015: 7_103_273_000,
       2016: 22_539_937_000, 2017: 23_943_611_000, 2018: 19_733_880_000,
       2019: 30_497_283_000, 2020: 2_892_450_000, 2021: 4_970_290_000,
       2022: 7_907_707_000, 2023: 29_274_737_000, 2024: 30_601_328_000,
       2025: 37_264_981_000, 2026: 7_710_096_000}

MESES = {'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
         'jul': 7, 'ago': 8, 'set': 9, 'sep': 9, 'oct': 10, 'nov': 11,
         'dic': 12}
MES_TOK = ('ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SET','SEP',
           'OCT','NOV','DIC')
RE_F = re.compile(r'\d{1,2}-[A-Za-z]{3}-\d{2}\b')
RE_M = re.compile(r'^\d{1,3}(,\d{3})+$')
RE_T = re.compile(r'^\d{1,3}\.\d{2,4}$')


def limpiar_num(s):
    """'1 00,000,000' -> 100000000.0 ; '6 .5570' -> 6.557 ; '-' -> None"""
    if s is None:
        return None
    s = re.sub(r'\s+', '', str(s).replace('\n', ' ')).replace(',', '')
    if s in ('', '-', '--'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fecha(s):
    if not s:
        return None
    m = re.match(r'^(\d{1,2})[/-]([A-Za-z]{3}|\d{2})[/-](\d{2,4})$',
                 str(s).strip())
    if not m:
        return None
    d, mth, y = m.groups()
    mth_n = int(mth) if mth.isdigit() else MESES.get(mth.lower())
    if not mth_n:
        return None
    y = int(y)
    if y < 100:          # en este documento todo año corto es 20xx (2001-2055)
        y += 2000
    try:
        return pd.Timestamp(y, mth_n, int(d))
    except ValueError:
        return None


def venc_de_denom(b):
    """'13OCT2024' / '04ENE2026A' / '12FEB2029E' -> Timestamp"""
    m = re.match(r'^(\d{2})([A-Z]{3})(\d{4})', str(b or ''))
    if not m:
        return None
    d, mes, y = m.groups()
    mn = MESES.get(mes.lower())
    return pd.Timestamp(int(y), mn, int(d)) if mn else None


def es_ruido(txt):
    return (('Tipo' in txt and 'Mes' in txt) or 'Denominación' in txt
            or 'Leyenda' in txt or re.match(r'^\s*\d+/ ', txt)
            or 'Emisión realizada' in txt or 'El monto total' in txt
            or 'Operación de Endeudamiento en el marco' in txt
            or 'EMISIONES INTERNAS' in txt)


# ============================================================================
# MOTOR 1: celdas de tabla de pdfplumber (partes separadas por \n)
# ============================================================================
def parse_v1(pdf_path):
    filas, totales = [], {}
    anio, ultimo = None, [None]

    with pdfplumber.open(pdf_path) as pdf:
        for pag in pdf.pages:
            for tabla in pag.extract_tables():
                if not tabla or max(len(r) for r in tabla) < 13:
                    continue
                ncol = max(len(r) for r in tabla)
                fp = [list(r) + [None] * (ncol - len(r)) for r in tabla]

                def frac_tipo(j):
                    vals = [str(f[j]).strip() for f in fp
                            if f[j] not in (None, '')]
                    vals = [v for v in vals
                            if not re.match(r'^(20\d\d|Total)', v)]
                    if not vals:
                        return 0.0
                    return sum(bool(re.match(r'^(OE|OAD|BDA)', v))
                               for v in vals) / len(vals)
                tipo_col = max(range(min(3, ncol)), key=frac_tipo)

                fc = [f[tipo_col:] for f in fp]
                nc2 = max(len(f) for f in fc)
                fc = [f + [None] * (nc2 - len(f)) for f in fc]
                keep = [j for j in range(nc2)
                        if any(c not in (None, '') for c in
                               (f[j] for f in fc))]
                ft = [[f[j] for j in keep] for f in fc]
                if len(keep) == 12:                 # falta Tasa Marginal
                    ft = [f[:10] + [None] + f[10:] for f in ft]
                elif len(keep) != 13:
                    ft = [(f + [None] * 13)[:13] for f in ft]

                bloque = []

                def procesar(b):
                    c0 = [str(c) if c not in (None, '') else ''
                          for c in (b[0] + [None] * 13)[:13]]
                    tipo_b, num_b, mes_b, norma_b = c0[0], c0[1], c0[2], c0[5]
                    fs_q, fe_q, sub = [], [], []
                    for fila in b:
                        cl = [str(c) if c not in (None, '') else ''
                              for c in (fila + [None] * 13)[:13]]
                        fs_q += [p.strip() for p in cl[3].split('\n')
                                 if p.strip()]
                        fe_q += [p.strip() for p in cl[4].split('\n')
                                 if p.strip()]
                        if cl[5].strip() and not norma_b:
                            norma_b = cl[5]
                        datos = [cl[j].split('\n') for j in range(6, 13)]
                        k = max(len(x) for x in datos)
                        for j in range(7):
                            x = datos[j]
                            if len(x) == 1 and k > 1:
                                datos[j] = x * k
                            elif len(x) < k:
                                datos[j] = x + [''] * (k - len(x))
                        for i in range(k):
                            sub.append([datos[j][i] for j in range(7)])
                    for i, sf in enumerate(sub):
                        if 'Bonos' not in sf[0]:
                            continue
                        fs = fs_q[min(i, len(fs_q) - 1)] if fs_q else None
                        fe = fe_q[min(i, len(fe_q) - 1)] if fe_q else None
                        f_s, f_e = fecha(fs), fecha(fe)
                        a_f = (f_s.year if f_s is not None else
                               (f_e.year if f_e is not None else None))
                        a = max(x for x in (anio, a_f) if x is not None)
                        ultimo[0] = a
                        filas.append(dict(zip(
                            ('Año', 'Tipo', 'Nº', 'Mes', 'f_sub', 'f_emi',
                             'Norma', 'denom', 'm_sub', 'm_col', 't_pro',
                             't_mar', 't_cup', 'f_ven', 'pag'),
                            (a, tipo_b, num_b, mes_b, fs, fe, norma_b,
                             sf[0], sf[1], sf[2], sf[3], sf[4], sf[5],
                             sf[6], pag.page_number))))

                def cerrar():
                    if bloque:
                        procesar(bloque)
                        bloque.clear()

                for fo, fila in zip(fp, ft):
                    txt = ' '.join(str(c) for c in fo if c)
                    nv = [str(c).strip() for c in fo if c not in (None, '')]
                    solo1 = nv[0] if len(nv) == 1 else None
                    m = re.match(r'^(20[0-2]\d)$', solo1 or '')
                    if m:
                        cerrar()
                        anio = int(m.group(1))
                        continue
                    if re.match(r'^\s*Total\b', txt):
                        cerrar()
                        nums = [n for n in re.findall(r'\d[\d,]*\d', txt)
                                if len(n.replace(',', '')) >= 7]
                        if nums and ultimo[0]:
                            totales[ultimo[0]] = float(
                                nums[-1].replace(',', ''))
                        continue
                    if es_ruido(txt):
                        cerrar()
                        continue
                    tiene = any('Bonos' in str(c) for c in fila if c)
                    es_cont = all(fila[j] in (None, '') for j in (0, 1, 2))
                    if es_cont and bloque:
                        bloque.append(fila)
                    elif tiene:
                        cerrar()
                        bloque.append(fila)
                    else:
                        cerrar()
                cerrar()
    return pd.DataFrame(filas), totales


# ============================================================================
# MOTOR 2: alineación de subfilas por coordenada vertical (Y) de cada palabra
# ============================================================================
def _lineas(pag, bbox):
    if bbox is None:
        return []
    try:
        reg = pag.crop(bbox)
    except ValueError:
        return []
    ws = sorted(reg.extract_words(), key=lambda w: (w['top'], w['x0']))
    out = []
    for w in ws:
        if out and abs(w['top'] - out[-1][0]) <= 2:
            out[-1] = (out[-1][0], out[-1][1] + ' ' + w['text'])
        else:
            out.append((w['top'], w['text']))
    return out


def parse_v4(pdf_path):
    filas, totales = [], {}
    anio, ultimo = None, [None]

    with pdfplumber.open(pdf_path) as pdf:
        for pag in pdf.pages:
            for tabla in pag.find_tables():
                if not tabla.rows or len(tabla.rows[0].cells) < 12:
                    continue
                ncol = len(tabla.rows[0].cells)
                grid = [[_lineas(pag, c) for c in r.cells]
                        for r in tabla.rows]

                def frac_tipo(j):
                    vals = [t for f in grid for _, t in f[j]]
                    vals = [v for v in vals if v and
                            not re.match(r'^(20\d\d|Total)', v)]
                    if not vals:
                        return 0
                    return sum(bool(re.match(r'^(OE|OAD|BDA)', v))
                               for v in vals) / len(vals)
                tipo_col = max(range(min(3, ncol)), key=frac_tipo)

                usadas = [j for j in range(tipo_col, ncol)
                          if any(f[j] for f in grid)]
                if len(usadas) == 12:
                    usadas = usadas[:10] + [None] + usadas[10:]
                if len(usadas) != 13:
                    continue

                bloque = []

                def procesar(b):
                    cols = [[] for _ in range(13)]
                    for f in b:
                        for j in range(13):
                            cols[j].extend(f[j])
                    for j in range(13):
                        cols[j].sort(key=lambda x: x[0])

                    def texto(j):
                        t = ' '.join(t for _, t in cols[j]).strip()
                        return t or None
                    tipo_v, num_v, mes_v = texto(0), texto(1), texto(2)
                    norma_v = texto(5)
                    fs_q = [t for _, tt in cols[3] for t in tt.split()]
                    fe_q = [t for _, tt in cols[4] for t in tt.split()]

                    ys = sorted(y for j in range(6, 13) for y, _ in cols[j])
                    anclas = []
                    for y in ys:
                        if not anclas or y - anclas[-1] > 4:
                            anclas.append(y)
                    if not anclas:
                        return

                    def asignar(j):
                        res = [None] * len(anclas)
                        for y, t in cols[j]:
                            i = min(range(len(anclas)),
                                    key=lambda k: abs(anclas[k] - y))
                            res[i] = t if res[i] is None else res[i] + ' ' + t
                        return res
                    D, MS, MC = asignar(6), asignar(7), asignar(8)
                    TP, TM, TC, FV = (asignar(9), asignar(10),
                                      asignar(11), asignar(12))
                    for i in range(len(anclas)):
                        if D[i] is None and i > 0:
                            D[i] = D[i - 1]
                    k_out = 0
                    for i in range(len(anclas)):
                        d = D[i]
                        if not d or 'Bonos' not in d:
                            continue
                        fs = fs_q[min(k_out, len(fs_q) - 1)] if fs_q else None
                        fe = fe_q[min(k_out, len(fe_q) - 1)] if fe_q else None
                        f_s, f_e = fecha(fs), fecha(fe)
                        a_f = (f_s.year if f_s is not None else
                               (f_e.year if f_e is not None else None))
                        a = max(x for x in (anio, a_f) if x is not None)
                        ultimo[0] = a
                        filas.append(dict(zip(
                            ('Año', 'Tipo', 'Nº', 'Mes', 'f_sub', 'f_emi',
                             'Norma', 'denom', 'm_sub', 'm_col', 't_pro',
                             't_mar', 't_cup', 'f_ven', 'pag'),
                            (a, tipo_v, num_v, mes_v, fs, fe, norma_v, d,
                             MS[i], MC[i], TP[i], TM[i], TC[i], FV[i],
                             pag.page_number))))
                        k_out += 1

                def cerrar():
                    if bloque:
                        procesar(bloque)
                        bloque.clear()

                for f_orig in grid:
                    f = [(f_orig[j] if j is not None else [])
                         for j in usadas]
                    txt = ' '.join(t for c in f_orig for _, t in c)
                    m = re.match(r'^\s*(20[0-2]\d)\s*$', txt)
                    if m:
                        cerrar()
                        anio = int(m.group(1))
                        continue
                    if re.match(r'^\s*Total\b', txt):
                        cerrar()
                        nums = [n for n in re.findall(r'\d[\d,]*\d', txt)
                                if len(n.replace(',', '')) >= 7]
                        if nums and ultimo[0]:
                            totales[ultimo[0]] = float(
                                nums[-1].replace(',', ''))
                        continue
                    if es_ruido(txt):
                        cerrar()
                        continue
                    tiene = 'Bonos' in txt
                    es_cont = not f[0] and not f[1] and not f[2]
                    if es_cont and bloque:
                        bloque.append(f)
                    elif tiene:
                        cerrar()
                        bloque.append(f)
                    else:
                        cerrar()
                cerrar()
    return pd.DataFrame(filas), totales


# ============================================================================
# MOTOR 3: líneas de texto de página (extract_text)
# ============================================================================
def _tokenizar(toks):
    out = []
    for t in toks:
        if out and re.match(r'^\d{1,2}$', out[-1]) and \
           (t.startswith('.') or (RE_M.match(t) and t.startswith('0'))
                or re.match(r'^\d{2},\d{3}', t)):
            out[-1] = out[-1] + t
        else:
            out.append(t)
    return out


def _parse_data(toks):
    toks = [t for t in toks if not re.match(r'^\d+/$', t)]
    fv = toks.pop() if toks and RE_F.match(toks[-1]) else None
    dinero, tasas = [], []
    for t in toks:
        if RE_M.match(t):
            dinero.append(t)
        elif RE_T.match(t):
            tasas.append(t)
        elif t == '-':
            (dinero if len(dinero) < 2 and not tasas else tasas).append(None)
    if len(dinero) >= 2:
        msub, mcol = dinero[0], dinero[1]
    elif len(dinero) == 1:
        msub, mcol = None, dinero[0]
    else:
        msub = mcol = None
    tasas = (tasas + [None] * 3)[:3]
    if sum(x is not None for x in tasas) == 2 and tasas[2] is None:
        tpro, tmar, tcup = tasas[0], None, tasas[1]
    else:
        tpro, tmar, tcup = tasas
    return msub, mcol, tpro, tmar, tcup, fv


def parse_v5(pdf_path):
    filas, totales = [], {}
    anio, ultimo = None, [None]

    with pdfplumber.open(pdf_path) as pdf:
        for pag in pdf.pages:
            texto = pag.extract_text() or ''
            meta, fechas_q, pendiente, i_f = {}, [], [], [0]

            def emitir(denom, datos, fs, fe):
                f_s, f_e = fecha(fs), fecha(fe)
                a_f = (f_s.year if f_s is not None else
                       (f_e.year if f_e is not None else None))
                a = max(x for x in (anio, a_f) if x is not None)
                ultimo[0] = a
                msub, mcol, tp, tm, tc, fv = datos
                filas.append(dict(zip(
                    ('Año', 'Tipo', 'Nº', 'Mes', 'f_sub', 'f_emi', 'Norma',
                     'denom', 'm_sub', 'm_col', 't_pro', 't_mar', 't_cup',
                     'f_ven', 'pag'),
                    (a, meta.get('t'), meta.get('n'), meta.get('m'), fs, fe,
                     meta.get('nl'), denom, msub, mcol, tp, tm, tc, fv,
                     pag.page_number))))

            def next_dates():
                if not fechas_q:
                    return None, None
                k = min(i_f[0], len(fechas_q) - 1)
                i_f[0] += 1
                return fechas_q[k]

            for ln in texto.split('\n'):
                s = ln.strip()
                if not s:
                    continue
                m = re.match(r'^(20[0-2]\d)$', s)
                if m:
                    anio = int(m.group(1))
                    continue
                if s.startswith('Total'):
                    nums = [n for n in re.findall(r'\d[\d,]*\d', s)
                            if len(n.replace(',', '')) >= 7]
                    if nums and ultimo[0]:
                        totales[ultimo[0]] = float(nums[-1].replace(',', ''))
                    continue
                if es_ruido(s) or s.startswith(('OE :', 'OAD:', 'BDA:')):
                    continue

                toks = _tokenizar(s.split())
                s = ' '.join(toks)
                mnew = re.match(r'^(OE(?:\s*/\s*OAD)?|OAD|BDA)\b', s)
                if mnew:
                    meta = {'t': mnew.group(1).replace(' ', '')}
                    resto = toks[1:]
                    if resto and re.match(r'^(\d{1,2}|FB\d|S\.E\.|\*\*|-)$',
                                          resto[0]):
                        if resto[0] not in ('**', '-'):
                            meta['n'] = resto[0]
                        resto = resto[1:]
                    if resto and resto[0] in MES_TOK:
                        meta['m'] = resto[0]
                    fechas_q, i_f[0], pendiente = [], 0, []

                if 'Bonos' in s:
                    pre = s.split('Bonos')[0]
                else:
                    pre = s
                    for k, t in enumerate(toks):
                        if RE_M.match(t) or RE_T.match(t):
                            pre = ' '.join(toks[:k])
                            break
                dts = RE_F.findall(pre)
                j = 0
                while j + 1 < len(dts):
                    fechas_q.append((dts[j], dts[j + 1]))
                    j += 2
                if j < len(dts):
                    fechas_q.append((dts[j], None))

                if 'Bonos' in s:
                    pre_n = re.sub(r'^(OE(\s*/\s*OAD)?|OAD|BDA)\s*', '', pre)
                    pre_n = re.sub(RE_F, '', pre_n)
                    pre_n = re.sub(r'^\s*(\d{1,2}|FB\d|S\.E\.|\*\*|-)\s+',
                                   '', pre_n)
                    pre_n = re.sub(
                        r'^\s*(' + '|'.join(MES_TOK) + r')\s+', '',
                        pre_n).strip()
                    if pre_n:
                        meta['nl'] = (meta.get('nl', '') + ' '
                                      + pre_n).strip()
                    seg = s[s.index('Bonos'):]
                    st = seg.split()
                    dt = st[:2]
                    k = 2
                    if len(st) > 2 and re.match(r'^\d{2}[A-Z]{3}\d{4}',
                                                st[2]):
                        dt.append(st[2])
                        k = 3
                    denom = ' '.join(dt)
                    post = st[k:]
                    if any(RE_M.match(t) or RE_T.match(t)
                           for t in post) or '-' in post:
                        fs, fe = next_dates()
                        emitir(denom, _parse_data(post), fs, fe)
                    else:
                        pendiente.append(denom)
                else:
                    if any(RE_M.match(t) or RE_T.match(t)
                           for t in toks) and RE_F.search(s):
                        if pendiente:
                            denom = pendiente.pop(0)
                        elif filas and filas[-1].get('denom'):
                            denom = filas[-1]['denom']
                        else:
                            continue
                        fs, fe = next_dates()
                        emitir(denom, _parse_data(toks), fs, fe)
    return pd.DataFrame(filas), totales


# ============================================================================
# Limpieza a columnas finales
# ============================================================================
FINAL = ['Año', 'Tipo', 'Nº', 'Mes', 'Fecha Subasta', 'Fecha Emisión',
         'Norma Legal', 'Denominación', 'Bono', 'Monto Subastado (S/)',
         'Monto Colocado (S/)', 'Tasa Promedio (%)', 'Tasa Marginal (%)',
         'Tasa Cupón (%)', 'Fecha Vencimiento']


def normalizar(df):
    if df.empty:
        return pd.DataFrame(columns=FINAL)
    df = df.copy()
    df['Monto Subastado (S/)'] = df['m_sub'].map(limpiar_num)
    df['Monto Colocado (S/)'] = df['m_col'].map(limpiar_num)
    df['Tasa Promedio (%)'] = df['t_pro'].map(limpiar_num)
    df['Tasa Marginal (%)'] = df['t_mar'].map(limpiar_num)
    df['Tasa Cupón (%)'] = df['t_cup'].map(limpiar_num)
    df['Fecha Subasta'] = df['f_sub'].map(fecha)
    df['Fecha Emisión'] = df['f_emi'].map(fecha)
    df['Denominación'] = (df['denom'].astype(str).str.replace('\n', ' ')
                          .str.replace(r'\s+', ' ', regex=True).str.strip())
    df['Bono'] = df['Denominación'].str.extract(r'Bonos Soberanos\s+(\S+)')
    # Vencimiento: columna del PDF; si falta, se deriva de la denominación
    fv = df['f_ven'].map(fecha)
    fv2 = df['Bono'].map(venc_de_denom)
    df['Fecha Vencimiento'] = fv.where(fv.notna(), fv2)
    for c, src in (('Tipo', 'Tipo'), ('Nº', 'Nº'), ('Mes', 'Mes'),
                   ('Norma Legal', 'Norma')):
        df[c] = (df[src].astype(str).str.replace('\n', ' ')
                 .str.replace(r'\s+', ' ', regex=True).str.strip()
                 .replace({'None': '', 'nan': ''}))
    df = df[df['Denominación'].str.contains('Bonos', na=False)]
    return df[FINAL].reset_index(drop=True)


# ============================================================================
# Ensamblaje y Excel
# ============================================================================
def elegir_pdf():
    """Abre una ventana para escoger el PDF y dónde guardar el Excel."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        print('tkinter no disponible; uso: python emisiones_mef_a_excel.py '
              '<archivo.pdf> [salida.xlsx]')
        sys.exit(1)
    raiz = tk.Tk()
    raiz.withdraw()
    raiz.attributes('-topmost', True)
    pdf = filedialog.askopenfilename(
        title='Escoge el PDF de emisiones del MEF',
        filetypes=[('PDF', '*.pdf'), ('Todos los archivos', '*.*')])
    if not pdf:
        print('No se escogió ningún archivo.')
        sys.exit(0)
    salida = filedialog.asksaveasfilename(
        title='Guardar Excel como...',
        defaultextension='.xlsx',
        initialfile='Emisiones_bonos_soberanos_MEF.xlsx',
        filetypes=[('Excel', '*.xlsx')])
    raiz.destroy()
    if not salida:
        salida = 'Emisiones_bonos_soberanos_MEF.xlsx'
    return pdf, salida


def main():
    # Con argumentos: python script.py archivo.pdf [salida.xlsx]
    # Sin argumentos: se abre la ventana para escoger el PDF
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
        salida = sys.argv[2] if len(sys.argv) > 2 else \
            'Emisiones_bonos_soberanos_MEF.xlsx'
    else:
        pdf_path, salida = elegir_pdf()
    print(f'PDF: {pdf_path}')

    print('Extrayendo con motor 1 (celdas de tabla)...')
    d1, _ = parse_v1(pdf_path)
    print('Extrayendo con motor 2 (coordenada vertical)...')
    d4, _ = parse_v4(pdf_path)
    print('Extrayendo con motor 3 (líneas de texto)...')
    d5, _ = parse_v5(pdf_path)
    d1, d4, d5 = normalizar(d1), normalizar(d4), normalizar(d5)

    piezas, reporte = [], []
    for a in sorted(REF):
        candidatos = [('celdas', d1), ('coordenada Y', d4),
                      ('líneas de texto', d5)]
        elegido, motor, dif = None, None, None
        for nombre, d in candidatos:
            g = d[d['Año'] == a]
            if g.empty:
                continue
            s = g['Monto Colocado (S/)'].sum()
            e = s - REF[a]
            if dif is None or abs(e) < abs(dif):
                elegido, motor, dif = g, nombre, e
            if e == 0:
                break
        if elegido is None:
            reporte.append((a, 0, 0, REF[a], None, 'SIN DATOS'))
            continue
        piezas.append(elegido)
        estado = 'OK' if dif == 0 else f'REVISAR (dif {dif:,.0f})'
        reporte.append((a, len(elegido),
                        elegido['Monto Colocado (S/)'].sum(),
                        REF[a], motor, estado))

    df = pd.concat(piezas, ignore_index=True)
    df = df.sort_values(['Año', 'Fecha Subasta', 'Nº'],
                        na_position='last').reset_index(drop=True)

    print(f"\n{'Año':>5} {'Filas':>6} {'Suma colocado':>18} "
          f"{'Total PDF':>18} {'Motor':>16}  Estado")
    for a, n, s, t, m, e in reporte:
        print(f'{a:>5} {n:>6} {s:>18,.0f} {t:>18,.0f} {str(m):>16}  {e}')

    escribir_excel(df, reporte, salida)
    print(f'\nListo: {salida}  ({len(df)} filas)')


def escribir_excel(df, reporte, ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    AZUL = '1F4E78'
    wb = Workbook()

    # ---------- Hoja Emisiones ----------
    ws = wb.active
    ws.title = 'Emisiones'
    fino = Side(style='thin', color='BFBFBF')
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)
    f_head = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    f_body = Font(name='Arial', size=10)

    for j, c in enumerate(df.columns, 1):
        cel = ws.cell(row=1, column=j, value=c)
        cel.font = f_head
        cel.fill = PatternFill('solid', fgColor=AZUL)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        cel.border = borde

    montos = ('Monto Subastado (S/)', 'Monto Colocado (S/)')
    tasas = ('Tasa Promedio (%)', 'Tasa Marginal (%)', 'Tasa Cupón (%)')
    fechas = ('Fecha Subasta', 'Fecha Emisión', 'Fecha Vencimiento')

    for i, fila in enumerate(df.itertuples(index=False), 2):
        for j, (col, v) in enumerate(zip(df.columns, fila), 1):
            if pd.isna(v):
                v = None
            elif col in fechas and v is not None:
                v = v.to_pydatetime()
            cel = ws.cell(row=i, column=j, value=v)
            cel.font = f_body
            cel.border = borde
            if col in montos:
                cel.number_format = '#,##0'
            elif col in tasas:
                cel.number_format = '0.0000'
            elif col in fechas:
                cel.number_format = 'dd/mm/yyyy'
                cel.alignment = Alignment(horizontal='center')
            elif col == 'Año':
                cel.number_format = '0'
                cel.alignment = Alignment(horizontal='center')
            elif col in ('Tipo', 'Nº', 'Mes', 'Bono'):
                cel.alignment = Alignment(horizontal='center')

    anchos = {'Año': 6, 'Tipo': 8, 'Nº': 6, 'Mes': 6, 'Fecha Subasta': 13,
              'Fecha Emisión': 13, 'Norma Legal': 42, 'Denominación': 30,
              'Bono': 12, 'Monto Subastado (S/)': 17,
              'Monto Colocado (S/)': 17, 'Tasa Promedio (%)': 12,
              'Tasa Marginal (%)': 12, 'Tasa Cupón (%)': 12,
              'Fecha Vencimiento': 14}
    for j, c in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(j)].width = anchos.get(c, 12)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(df.columns))}1'

    # ---------- Hoja Validación ----------
    wv = wb.create_sheet('Validación')
    cab = ['Año', 'Filas', 'Suma Monto Colocado (S/)',
           'Total según PDF (S/)', 'Diferencia (S/)', 'Motor usado',
           'Estado']
    for j, c in enumerate(cab, 1):
        cel = wv.cell(row=1, column=j, value=c)
        cel.font = f_head
        cel.fill = PatternFill('solid', fgColor=AZUL)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        cel.border = borde
    for i, (a, n, s, t, m, e) in enumerate(reporte, 2):
        vals = [a, n, s, t, s - t, m, e]
        for j, v in enumerate(vals, 1):
            cel = wv.cell(row=i, column=j, value=v)
            cel.font = f_body
            cel.border = borde
            if j in (3, 4, 5):
                cel.number_format = '#,##0'
            if j in (1, 2, 6, 7):
                cel.alignment = Alignment(horizontal='center')
            if e != 'OK' and j == 7:
                cel.font = Font(name='Arial', size=10, bold=True,
                                color='C00000')
    n = len(reporte) + 3
    nota = ('Validación: suma anual de "Monto Colocado" contra la fila '
            '"Total" impresa en el PDF del MEF. Para cada año se usó el '
            'motor de extracción con cuadre exacto; los años marcados '
            'REVISAR tienen la diferencia indicada y conviene contrastar '
            'sus filas contra el PDF.')
    wv.cell(row=n, column=1, value=nota).font = Font(name='Arial', size=9,
                                                     italic=True)
    for j, w in enumerate((6, 8, 22, 22, 16, 16, 24), 1):
        wv.column_dimensions[get_column_letter(j)].width = w
    wv.freeze_panes = 'A2'

    wb.save(ruta)


if __name__ == '__main__':
    main()
